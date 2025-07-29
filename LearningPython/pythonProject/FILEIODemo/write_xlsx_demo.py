from openpyxl import Workbook
wb = Workbook()
ws = wb.active
#ws['A1'] = "RCV Academy"

testdata = [["Name","city"],["Prathiksha","paris"],["Soundarya","singapore"],["sam","London"]]
for data in testdata:
    ws.append(data)

# for i in range(4,6):
#     for j in range(2, 5):
#         ws.cell(row=i , column = j).value = i+j
wb.save("demoexcel.xlsx")

# from openpyxl import Workbook
#
# wb = Workbook()
#
# ws = wb.active
# #
# # # ws['D1'] = "Prathiksha"
# # testdata = [['Name', 'City'], ['Prathiksha', 'Mysore'], ['Soundarya', 'Bengaluru'], ['Apeksha', 'Chennai']]
# # for data in testdata:
# #     ws.append(data)
#
# for i in range(1,6):
#     for j in range(1, 5):
#         ws.cell(row=i, column=j).value = i+j
# wb.save("demoexcel1.xlsx")


#Write to Excel using openpyxl (cell by cell)

from openpyxl import Workbook
# from openpyxl.workbook import Workbook
#
# wb = Workbook()
#
# ws = wb.active
#
# ws['A1'] = "Prathiksha"
# ws['B1'] = 25
# ws['A2'] = "Soundarya"
# ws['B2'] = 23
# ws['A3'] = "Sudha"
# ws['B3'] = 42
#
# wb.save("employee.xlsx")

#Write list of rows using openpyxl.append()


# from openpyxl import Workbook
#
# data = [
#     ["Name",'Age'],
#     ["Prathiksha", 23],
#     ["Soundarya",21],
#     ["bhavana",12],
# ]
# wb = Workbook()
# ws = wb.active
#
# for row in data:
#     ws.append(row)
#     wb.save("people.xlsx")

# # #Write using pandas (easy for tables)
# import pandas as pd
#
# data = {
#     "Product": ["Apple", "Banana", "Orange"],
#     "Price": [1.2, 0.5, 0.8],
#     "Stock": [100, 150, 200]
# }
#
# df = pd.DataFrame(data)
# df.to_excel("products.xlsx", index=False)

# import pandas as pd
# data = {
#      "Product": ["Sugar", "Salt", "Coffee"],
#      "Price": [20, 90, 33]
#  }
# df = pd.DataFrame(data)
# df.to_excel("productreview.xlsx",index=False)

#4 Write multiple sheets in one Excel file (pandas)
#
# import pandas as pd
#
# df1 = pd.DataFrame({"Name": ["Alice", "Bob"], "Age": [25, 30]})
# df2 = pd.DataFrame({"City": ["NY", "LA"], "Population": [8000000, 4000000]})
#
# with pd.ExcelWriter("multi_sheet.xlsx", engine="openpyxl") as writer:
#     df1.to_excel(writer, sheet_name="People", index=False)
#     df2.to_excel(writer, sheet_name="Cities", index=False)
#
# import pandas as pd
#
# df1 = pd.DataFrame({"Name":["Prathiksha","Soundarya"], "Age":[23,29]})
# df2 = pd.DataFrame({"RCA":["Skills","Performance"],"Score":[89,90]})
#
# with pd.ExcelWriter("multi_sheet1.xlsx",engine="openpyxl")as writer:
#     df1.to_excel(writer,sheet_name="Employee",index=False)
#     df2.to_excel(writer,sheet_name="Yearly performance",index=False)

 #5 Write a multiplication table using openpyxl

from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "Multiplication"

for i in range(1, 11):
        for j in range(1, 11):
            ws.cell(row=i, column=j).value = i * j

wb.save("multiplication_table.xlsx")

