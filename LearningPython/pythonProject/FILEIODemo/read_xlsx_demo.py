# from openpyxl import load_workbook
#
# wb = load_workbook(filename='demoexcel.xlsx')
# sh = wb['Sheet']
#
# row_ct = sh.max_row
# col_ct = sh.max_column
#
# for i in range(1,row_ct+1):
#     for j in range(1, col_ct+1):
#         print(sh.cell(row=i , column = j).value, end = ' ')
#         print('\n')

#Read an Excel File
# import pandas as pd
# df = pd.read_excel("Productreview.xlsx")
# print(df)


# 4
from openpyxl import load_workbook
from pandas.io.sas.sas_constants import row_count_offset_multiplier

Wb = load_workbook("multi_sheet1.xlsx")
sh = Wb['Yearly performance']

row_ct = sh.max_row
col_ct = sh.max_column
for i in range(1, row_ct+1):
    for j in range(1, col_ct+1):
        print(sh.cell(row = i, column = j).value)

#
# print(row_ct)
# print(col_ct)
# print(sh['A3'].value)
# print(Wb['Employee']['A2'].value)


# print(row_ct)
# print(col_ct)
# print(sh['A1'].value)
# print(wb['Sheet']['A2'].value)
# print(sh.cell(2, 2).value)