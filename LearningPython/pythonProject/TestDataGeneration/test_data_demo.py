from faker import Faker
from openpyxl import Workbook

# wb = Workbook()
# ws = wb.active
# fake_data = Faker(['hi_IN','en_US'])
#
# for i in range(1,11):
#     ws.cell(row=i , column=1).value = fake_data.name()
#     ws.cell(row=i, column=2).value = fake_data.email()
#     ws.cell(row=i, column=3).value = fake_data.phone_number()
# wb.save("textdata.xlsx")
# # from multiprocessing.pool import worker

# from faker import Faker
# from openpyxl import Workbook
# wb = Workbook()
# ws = wb.active
# fake_data = Faker()
#
# for i in range(1,11):
#     for j in range(1,4):
#      ws.cell(row=i, column=1).value =fake_data.name()
#      ws.cell(row=i, column=2).value = fake_data.email()
#      ws.cell(row=i, column=3).value = fake_data.date_of_birth()
# wb.save("testdata.xlsx")
#
# fake_data = Faker()
# print(fake_data.name())
# print(fake_data.email())
# print(fake_data.phone_number())
# print(fake_data.name())
# print(fake_data.email())
# print(fake_data.address())

#Generate Fake Data into a List of Dictionaries

from faker import Faker

fake = Faker()
data = []

for i in range(3):
    person = {
        "name": fake.name(),
        "email": fake.email(),
        "job": fake.job(),
        "city": fake.city()
    }
    data.append(person)

for d in data:
    print(d)

#Write Fake Data to a CSV File
#
# import csv
# from faker import Faker
#
# fake = Faker()
#
# with (open("fake_people.csv", "w", newline="") as f):
#     writer = csv.writer(f)
#     writer.writerow(["Name", "Email", "Phone", "City"])
#
#     for i in range(10):
#         writer.writerow([fake.name(), fake.email(), fake.phone_number(), fake.city()])

